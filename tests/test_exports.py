"""Tests pour le module d'export Phase 5c."""

from __future__ import annotations

from pathlib import Path

import pytest

from db.connection import set_db_file
from db.migrations.runner import run_migrations
from db.models.evenements import (
    add_benevole,
    add_depense,
    add_evenement,
    add_tarif,
    add_vente,
    add_vente_ligne,
    set_parametre,
)
from db.models.tombola import add_lot


@pytest.fixture(autouse=True)
def setup_db(tmp_db):
    set_db_file(str(tmp_db))
    run_migrations()
    yield
    set_db_file("")


def _creer_evenement(nom: str = "Fête de l'école") -> int:
    return add_evenement(nom, "Fête", "Une belle fête", "2026-06-01", "2026-06-01", "termine", 500.0)


def _creer_evenement_complet() -> int:
    evt_id = _creer_evenement()

    # Tarif
    tarif_id = add_tarif(evt_id, "Adulte", 5.0, 0, 1)

    # Vente
    vente_id = add_vente(evt_id, "2026-06-01", "sur_place", "especes", None, 10.0, 0.0, 10.0, None)
    add_vente_ligne(vente_id, tarif_id, 2, 5.0)

    # Dépense
    add_depense(evt_id, "Décoration", 30.0, "2026-06-01", "Décor", None, "especes", None)

    # Bénévole
    add_benevole(evt_id, None, "Dupont", "Marie", "Accueil", "09:00", "17:00", "confirme")

    return evt_id


# ── Tests ConfigAsso ─────────────────────────────────────────────────────────


def test_get_config_asso_defaut():
    """Si paramètres vides → ConfigAsso avec valeurs vides (pas d'erreur)."""
    from core.exports import get_config_asso

    config = get_config_asso()
    assert config.nom == ""
    assert config.adresse == ""
    assert config.telephone == ""
    assert config.email == ""
    assert config.logo_path == ""


def test_get_config_asso_avec_valeurs():
    """Après set_parametre → ConfigAsso avec les bonnes valeurs."""
    set_parametre("asso_nom", "Les Interactifs")
    set_parametre("asso_adresse", "12 rue de l'École")
    set_parametre("asso_telephone", "0123456789")
    set_parametre("asso_email", "contact@interactifs.fr")
    set_parametre("asso_logo_path", "/chemin/logo.png")

    from core.exports import get_config_asso

    config = get_config_asso()
    assert config.nom == "Les Interactifs"
    assert config.adresse == "12 rue de l'École"
    assert config.telephone == "0123456789"
    assert config.email == "contact@interactifs.fr"
    assert config.logo_path == "/chemin/logo.png"


# ── Tests nommage de fichier ─────────────────────────────────────────────────


def test_nommage_fichier_export():
    """Vérifie le slug du nom de fichier."""
    from core.exports import generer_nom_fichier, slugifier_nom

    assert slugifier_nom("Fête de l'école") == "fete_de_lecole"
    assert slugifier_nom("Marché de Noël 2026") == "marche_de_noel_2026"
    assert slugifier_nom("Kermesse !!! 2026") == "kermesse_2026"
    assert slugifier_nom("") == "evenement"

    nom_fichier = generer_nom_fichier("Fête de l'école", "20260601", "pdf")
    assert nom_fichier == "bilan_fete_de_lecole_20260601.pdf"

    nom_xlsx = generer_nom_fichier("Fête de l'école", "20260601", "xlsx")
    assert nom_xlsx == "bilan_fete_de_lecole_20260601.xlsx"


# ── Tests export PDF ─────────────────────────────────────────────────────────


def test_export_pdf_evenement_vide(tmp_path: Path):
    """Événement sans données → PDF généré sans erreur."""
    evt_id = _creer_evenement()
    chemin = str(tmp_path / "bilan_vide.pdf")

    from core.exports import export_bilan_evenement_pdf

    result = export_bilan_evenement_pdf(evt_id, chemin)
    assert result is True
    assert Path(chemin).exists()
    assert Path(chemin).stat().st_size > 0


def test_export_pdf_evenement_complet(tmp_path: Path):
    """Événement avec billetterie + dépenses + bénévoles → PDF valide."""
    evt_id = _creer_evenement_complet()
    chemin = str(tmp_path / "bilan_complet.pdf")

    from core.exports import export_bilan_evenement_pdf

    result = export_bilan_evenement_pdf(evt_id, chemin)
    assert result is True
    assert Path(chemin).exists()
    assert Path(chemin).stat().st_size > 0


def test_export_pdf_evenement_introuvable(tmp_path: Path):
    """Événement inexistant → retourne False sans lever d'exception."""
    from core.exports import export_bilan_evenement_pdf

    result = export_bilan_evenement_pdf(999999, str(tmp_path / "inexistant.pdf"))
    assert result is False


def test_export_pv_tirage(tmp_path: Path):
    """PV avec lots → PDF généré."""
    evt_id = _creer_evenement()
    add_lot(evt_id, 1, "Panier garni", 30.0, 30.0, "achete", None, None, None)
    add_lot(evt_id, 2, "Bon cadeau", 50.0, 50.0, "sponsorise", None, "MagasinX", None)

    chemin = str(tmp_path / "pv_tirage.pdf")
    from core.exports import export_pv_tirage_pdf

    result = export_pv_tirage_pdf(evt_id, chemin)
    assert result is True
    assert Path(chemin).exists()
    assert Path(chemin).stat().st_size > 0


def test_export_liste_benevoles_pdf(tmp_path: Path):
    """Liste bénévoles → PDF généré."""
    evt_id = _creer_evenement_complet()
    chemin = str(tmp_path / "benevoles.pdf")

    from core.exports import export_liste_benevoles_pdf

    result = export_liste_benevoles_pdf(evt_id, chemin)
    assert result is True
    assert Path(chemin).exists()
    assert Path(chemin).stat().st_size > 0


def test_export_liste_benevoles_pdf_vide(tmp_path: Path):
    """Événement sans bénévoles → PDF généré sans erreur."""
    evt_id = _creer_evenement()
    chemin = str(tmp_path / "benevoles_vide.pdf")

    from core.exports import export_liste_benevoles_pdf

    result = export_liste_benevoles_pdf(evt_id, chemin)
    assert result is True
    assert Path(chemin).exists()


# ── Tests export Excel ───────────────────────────────────────────────────────


def test_export_excel_structure(tmp_path: Path):
    """Vérifie que les onglets attendus sont présents dans le classeur Excel."""
    from openpyxl import load_workbook

    evt_id = _creer_evenement_complet()
    chemin = str(tmp_path / "bilan.xlsx")

    from core.exports import export_bilan_evenement_excel

    result = export_bilan_evenement_excel(evt_id, chemin)
    assert result is True

    wb = load_workbook(chemin)
    onglets = wb.sheetnames
    assert "Résumé" in onglets
    assert "Billetterie" in onglets
    assert "Dépenses" in onglets
    assert "Bénévoles" in onglets


def test_export_excel_evenement_vide(tmp_path: Path):
    """Événement sans données → Excel généré sans erreur, onglets de base présents."""
    from openpyxl import load_workbook

    evt_id = _creer_evenement()
    chemin = str(tmp_path / "bilan_vide.xlsx")

    from core.exports import export_bilan_evenement_excel

    result = export_bilan_evenement_excel(evt_id, chemin)
    assert result is True

    wb = load_workbook(chemin)
    assert "Résumé" in wb.sheetnames


def test_export_excel_avec_tombola(tmp_path: Path):
    """Événement avec tombola → onglets Tombola présents."""
    from openpyxl import load_workbook

    evt_id = _creer_evenement()
    add_lot(evt_id, 1, "Lot 1", 20.0, 20.0, "achete", None, None, None)

    chemin = str(tmp_path / "bilan_tombola.xlsx")
    from core.exports import export_bilan_evenement_excel

    result = export_bilan_evenement_excel(evt_id, chemin)
    assert result is True

    wb = load_workbook(chemin)
    assert "Tombola_Lots" in wb.sheetnames


def test_export_liste_benevoles_excel(tmp_path: Path):
    """Liste bénévoles → Excel généré."""
    evt_id = _creer_evenement_complet()
    chemin = str(tmp_path / "benevoles.xlsx")

    from core.exports import export_liste_benevoles_excel

    result = export_liste_benevoles_excel(evt_id, chemin)
    assert result is True
    assert Path(chemin).exists()

    from openpyxl import load_workbook
    wb = load_workbook(chemin)
    assert "Bénévoles" in wb.sheetnames


def test_export_excel_introuvable(tmp_path: Path):
    """Événement inexistant → retourne False sans lever d'exception."""
    from core.exports import export_bilan_evenement_excel

    result = export_bilan_evenement_excel(999999, str(tmp_path / "inexistant.xlsx"))
    assert result is False


# ── Tests logo manquant ──────────────────────────────────────────────────────


def test_export_pdf_sans_logo(tmp_path: Path):
    """Logo inexistant → PDF généré sans erreur (sans logo)."""
    set_parametre("asso_nom", "Mon Asso")
    set_parametre("asso_logo_path", "/chemin/inexistant/logo.png")

    evt_id = _creer_evenement()
    chemin = str(tmp_path / "bilan_sans_logo.pdf")

    from core.exports import export_bilan_evenement_pdf

    result = export_bilan_evenement_pdf(evt_id, chemin)
    assert result is True
    assert Path(chemin).exists()
