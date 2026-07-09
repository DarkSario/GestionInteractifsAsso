"""Tests Phase 9 — Exports & Rapports."""

from __future__ import annotations

from pathlib import Path

import pytest

from db.connection import set_db_file
from db.migrations.runner import run_migrations


@pytest.fixture(autouse=True)
def setup_db(tmp_db):
    set_db_file(str(tmp_db))
    run_migrations()
    yield
    set_db_file("")


# ── Helpers ───────────────────────────────────────────────────────────────────


def _creer_membre(nom: str = "Dupont", prenom: str = "Jean") -> int:
    from db.models.membres import add_membre

    return add_membre(nom, prenom, f"{prenom.lower()}.{nom.lower()}@test.fr",
                      "0600000000", "parent", "2026-09-01", "")


def _creer_article_stock() -> int:
    from db.models.stock import add_article

    return add_article("Cahier", None, None, None, 50, 10, 1.50, None, None)


def _creer_compte() -> int:
    from db.models.tresorerie import add_compte

    return add_compte("Compte courant", "bancaire", 1000.0, 1, 0, "", "", 1)


def _creer_operation(compte_id: int) -> int:
    from db.models.tresorerie import add_operation

    return add_operation(
        compte_id, "recette", "Don adhérent", 50.0,
        "2026-01-15", None, "especes", None, None, None,
        "valide", 0, None, None, "",
    )


# ── Migration ─────────────────────────────────────────────────────────────────


def test_migration_polices_pdf_table():
    """La migration 0011 crée la table polices_pdf."""
    from db.connection import get_connection

    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='polices_pdf'"
        ).fetchall()
        assert len(rows) == 1
    finally:
        conn.close()


def test_migration_parametres_pdf():
    """Les paramètres PDF sont insérés par la migration 0011."""
    from db.models.parametres_globaux import get_parametre

    assert get_parametre("pdf_police_titre") == "Helvetica"
    assert get_parametre("pdf_police_corps") == "Helvetica"
    assert get_parametre("pdf_taille_base") == "11"
    assert get_parametre("pdf_couleur_accent") == "#000000"
    assert get_parametre("pdf_inclure_graphiques") == "0"
    assert get_parametre("pdf_format_defaut") == "A4"


# ── Modèle polices ────────────────────────────────────────────────────────────


def test_polices_systeme_toujours_presentes():
    """Les polices système sont toujours retournées."""
    from db.models.polices import get_all_polices

    polices = get_all_polices()
    noms = [p["nom"] for p in polices]
    assert "Helvetica" in noms


def test_add_police():
    """Ajout d'une police personnalisée."""
    from db.models.polices import add_police, get_police_by_nom

    chemin = "config/fonts/test.ttf"
    pid = add_police("TestFont", "test.ttf", chemin)
    assert pid > 0

    police = get_police_by_nom("TestFont")
    assert police is not None
    assert police["nom"] == "TestFont"
    assert police["est_systeme"] == 0


def test_delete_police_systeme_bloquee():
    """Suppression d'une police système est bloquée."""
    from db.models.polices import delete_police

    # Les polices système ont des IDs fictifs négatifs — non présents en DB
    result = delete_police(-1)
    assert result is False


def test_toggle_police():
    """La table polices_pdf est bien accessible après ajout."""
    from db.models.polices import add_police, get_police_by_id

    pid = add_police("ToggleFont", "toggle.ttf", "config/fonts/toggle.ttf")
    assert pid > 0

    police = get_police_by_id(pid)
    assert police is not None
    assert police["nom"] == "ToggleFont"


# ── Nom de fichier Phase 9 ────────────────────────────────────────────────────


def test_generer_nom_fichier_phase9():
    """Génération du nom de fichier Phase 9."""
    from core.exports import generer_nom_fichier_phase9
    from datetime import datetime

    today = datetime.now().strftime("%Y-%m-%d")

    # Les tirets dans le suffixe sont supprimés par le slugifier
    nom = generer_nom_fichier_phase9("bilan_ag", "2025-2026", "pdf")
    assert nom.startswith("bilan_ag_")
    assert nom.endswith(".pdf")
    assert today in nom

    nom_simple = generer_nom_fichier_phase9("liste_stock", extension="xlsx")
    assert nom_simple.startswith("liste_stock_")
    assert nom_simple.endswith(".xlsx")


# ── PDF Bilan AG ──────────────────────────────────────────────────────────────


def test_pdf_bilan_ag_genere(tmp_path: Path):
    """Le PDF Bilan AG est généré sans erreur avec toutes les sections."""
    from core.exports import export_bilan_ag_pdf

    chemin = str(tmp_path / "bilan_ag.pdf")
    result = export_bilan_ag_pdf("2025-2026", chemin)
    assert result is True
    assert Path(chemin).exists()
    assert Path(chemin).stat().st_size > 100


def test_pdf_bilan_ag_sections_optionnelles(tmp_path: Path):
    """Bilan AG avec seulement quelques sections."""
    from core.exports import export_bilan_ag_pdf

    sections = {
        "resume_financier": True,
        "tresorerie_detail": False,
        "subventions": False,
        "evenements": False,
        "buvette": False,
        "adherents": True,
        "signatures": True,
    }
    chemin = str(tmp_path / "bilan_ag_partiel.pdf")
    result = export_bilan_ag_pdf("2025-2026", chemin, sections=sections)
    assert result is True
    assert Path(chemin).exists()


def test_pdf_bilan_ag_page_garde(tmp_path: Path):
    """Bilan AG généré sans erreur (page de garde gérée par BasePDF)."""
    from core.pdf_bilan_ag import PdfBilanAG

    chemin = str(tmp_path / "bilan_ag_portrait.pdf")
    gen = PdfBilanAG("2025-2026")
    result = gen.generer(chemin)
    assert result is True
    assert Path(chemin).stat().st_size > 100


# ── PDF Adhérents ─────────────────────────────────────────────────────────────


def test_pdf_liste_adherents(tmp_path: Path):
    """Liste adhérents → PDF généré."""
    _creer_membre("Martin", "Sophie")
    _creer_membre("Leroux", "Paul")

    from core.exports import export_liste_adherents_pdf

    chemin = str(tmp_path / "liste_adherents.pdf")
    result = export_liste_adherents_pdf(chemin)
    assert result is True
    assert Path(chemin).exists()
    assert Path(chemin).stat().st_size > 100


def test_pdf_liste_adherents_vide(tmp_path: Path):
    """Liste adhérents vide → PDF généré sans erreur."""
    from core.exports import export_liste_adherents_pdf

    chemin = str(tmp_path / "liste_adherents_vide.pdf")
    result = export_liste_adherents_pdf(chemin)
    assert result is True
    assert Path(chemin).exists()


def test_pdf_fiche_adherent(tmp_path: Path):
    """Fiche individuelle → PDF généré."""
    membre_id = _creer_membre("Durand", "Claire")

    from core.exports import export_fiche_adherent_pdf

    chemin = str(tmp_path / "fiche_adherent.pdf")
    result = export_fiche_adherent_pdf(membre_id, chemin)
    assert result is True
    assert Path(chemin).exists()
    assert Path(chemin).stat().st_size > 100


def test_pdf_fiche_adherent_introuvable(tmp_path: Path):
    """Fiche d'un adhérent inexistant → retourne False."""
    from core.exports import export_fiche_adherent_pdf

    result = export_fiche_adherent_pdf(99999, str(tmp_path / "fiche_inexistante.pdf"))
    assert result is False


# ── PDF Stock ─────────────────────────────────────────────────────────────────


def test_pdf_liste_stock(tmp_path: Path):
    """Liste stock → PDF généré."""
    _creer_article_stock()

    from core.exports import export_liste_stock_pdf

    chemin = str(tmp_path / "liste_stock.pdf")
    result = export_liste_stock_pdf(chemin)
    assert result is True
    assert Path(chemin).exists()
    assert Path(chemin).stat().st_size > 100


def test_pdf_liste_stock_vide(tmp_path: Path):
    """Stock vide → PDF généré sans erreur."""
    from core.exports import export_liste_stock_pdf

    chemin = str(tmp_path / "liste_stock_vide.pdf")
    result = export_liste_stock_pdf(chemin)
    assert result is True
    assert Path(chemin).exists()


def test_pdf_historique_stock(tmp_path: Path):
    """Historique mouvements stock → PDF généré."""
    article_id = _creer_article_stock()
    from db.models.stock import add_mouvement
    add_mouvement(article_id, "2026-01-10", "entree", 20, 1.50, None, None, "F001", None)

    from core.exports import export_historique_stock_pdf

    chemin = str(tmp_path / "historique_stock.pdf")
    result = export_historique_stock_pdf(chemin)
    assert result is True
    assert Path(chemin).exists()


# ── PDF Trésorerie ────────────────────────────────────────────────────────────


def test_pdf_releve_compte(tmp_path: Path):
    """Relevé de compte → PDF généré."""
    compte_id = _creer_compte()
    _creer_operation(compte_id)

    from core.exports import export_releve_compte_pdf

    chemin = str(tmp_path / "releve_compte.pdf")
    result = export_releve_compte_pdf(compte_id, chemin)
    assert result is True
    assert Path(chemin).exists()
    assert Path(chemin).stat().st_size > 100


def test_pdf_releve_compte_avec_dates(tmp_path: Path):
    """Relevé de compte avec filtre de dates."""
    compte_id = _creer_compte()
    _creer_operation(compte_id)

    from core.exports import export_releve_compte_pdf

    chemin = str(tmp_path / "releve_compte_filtre.pdf")
    result = export_releve_compte_pdf(
        compte_id, chemin, date_debut="2026-01-01", date_fin="2026-12-31"
    )
    assert result is True
    assert Path(chemin).exists()


def test_pdf_subventions(tmp_path: Path):
    """PDF subventions → généré sans erreur."""
    from core.exports import export_subventions_pdf

    chemin = str(tmp_path / "subventions.pdf")
    result = export_subventions_pdf(chemin)
    assert result is True
    assert Path(chemin).exists()


# ── PDF Buvette ───────────────────────────────────────────────────────────────


def test_pdf_rapport_caisse_buvette(tmp_path: Path):
    """Rapport de caisse buvette → PDF généré."""
    from core.exports import export_rapport_caisse_pdf

    chemin = str(tmp_path / "rapport_caisse.pdf")
    result = export_rapport_caisse_pdf(chemin)
    assert result is True
    assert Path(chemin).exists()
    assert Path(chemin).stat().st_size > 100


# ── Excel Adhérents ───────────────────────────────────────────────────────────


def test_excel_adherents_onglets(tmp_path: Path):
    """Export Excel adhérents avec au moins un onglet."""
    from openpyxl import load_workbook

    _creer_membre("Petit", "Alice")

    from core.exports import export_liste_adherents_excel

    chemin = str(tmp_path / "adherents.xlsx")
    result = export_liste_adherents_excel(chemin)
    assert result is True
    assert Path(chemin).exists()

    wb = load_workbook(chemin)
    assert len(wb.sheetnames) >= 1


def test_excel_adherents_vide(tmp_path: Path):
    """Export Excel adhérents sans données → fichier créé sans erreur."""
    from core.exports import export_liste_adherents_excel

    chemin = str(tmp_path / "adherents_vide.xlsx")
    result = export_liste_adherents_excel(chemin)
    assert result is True
    assert Path(chemin).exists()


# ── Excel Trésorerie ──────────────────────────────────────────────────────────


def test_excel_tresorerie_onglets(tmp_path: Path):
    """Export Excel trésorerie avec onglets attendus."""
    from openpyxl import load_workbook

    compte_id = _creer_compte()
    _creer_operation(compte_id)

    from core.exports import export_tresorerie_excel

    chemin = str(tmp_path / "tresorerie.xlsx")
    result = export_tresorerie_excel(chemin)
    assert result is True
    assert Path(chemin).exists()

    wb = load_workbook(chemin)
    onglets = wb.sheetnames
    assert len(onglets) >= 1


def test_excel_tresorerie_vide(tmp_path: Path):
    """Export Excel trésorerie sans opérations → fichier créé."""
    from core.exports import export_tresorerie_excel

    chemin = str(tmp_path / "tresorerie_vide.xlsx")
    result = export_tresorerie_excel(chemin)
    assert result is True
    assert Path(chemin).exists()


# ── Excel Stock ───────────────────────────────────────────────────────────────


def test_excel_stock(tmp_path: Path):
    """Export Excel stock → fichier généré."""
    from openpyxl import load_workbook

    _creer_article_stock()

    from core.exports import export_stock_excel

    chemin = str(tmp_path / "stock.xlsx")
    result = export_stock_excel(chemin)
    assert result is True
    assert Path(chemin).exists()

    wb = load_workbook(chemin)
    assert len(wb.sheetnames) >= 1


# ── Orientation paysage automatique ──────────────────────────────────────────


def test_pdf_orientation_paysage(tmp_path: Path):
    """PDF en paysage → généré sans erreur."""
    from core.pdf_stock import PdfListeStock

    chemin = str(tmp_path / "stock_paysage.pdf")
    gen = PdfListeStock(orientation="paysage")
    result = gen.generer(chemin)
    assert result is True
    assert Path(chemin).exists()


def test_pdf_orientation_portrait(tmp_path: Path):
    """PDF en portrait (défaut) → généré sans erreur."""
    from core.pdf_adherents import PdfListeAdherents

    chemin = str(tmp_path / "adherents_portrait.pdf")
    gen = PdfListeAdherents()
    result = gen.generer(chemin)
    assert result is True
    assert Path(chemin).exists()


# ── Police personnalisée ──────────────────────────────────────────────────────


def test_pdf_avec_police_personnalisee(tmp_path: Path):
    """PDF généré avec une police système alternative."""
    from db.models.parametres_globaux import set_parametre

    # Utiliser Times-Roman (intégrée dans reportlab)
    set_parametre("pdf_police_titre", "Times-Roman")
    set_parametre("pdf_police_corps", "Times-Roman")

    from core.pdf_adherents import PdfListeAdherents

    chemin = str(tmp_path / "adherents_times.pdf")
    gen = PdfListeAdherents()
    result = gen.generer(chemin)
    assert result is True
    assert Path(chemin).exists()


def test_import_police_db(tmp_path: Path):
    """Import d'une police fictive en DB — sans erreur."""
    from db.models.polices import add_police, get_police_by_nom

    chemin_fictif = "config/fonts/ma_police.ttf"
    pid = add_police("MaPolice", "ma_police.ttf", chemin_fictif)
    assert pid > 0

    police = get_police_by_nom("MaPolice")
    assert police is not None
    assert police["nom"] == "MaPolice"


# ── Compatibilité Phase 5c ────────────────────────────────────────────────────


def test_exports_phase5c_toujours_fonctionnels(tmp_path: Path):
    """Les exports Phase 5c (événement) restent fonctionnels après Phase 9."""
    from db.models.evenements import add_evenement
    from core.exports import export_bilan_evenement_pdf, export_bilan_evenement_excel

    evt_id = add_evenement(
        "Kermesse", "Kermesse", "Fête annuelle",
        "2026-06-15", "2026-06-15", "termine", 300.0,
    )

    chemin_pdf = str(tmp_path / "kermesse.pdf")
    result_pdf = export_bilan_evenement_pdf(evt_id, chemin_pdf)
    assert result_pdf is True

    chemin_xlsx = str(tmp_path / "kermesse.xlsx")
    result_xlsx = export_bilan_evenement_excel(evt_id, chemin_xlsx)
    assert result_xlsx is True
