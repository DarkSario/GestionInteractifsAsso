"""Export Excel multi-onglets pour le Dossier de Subvention — Phase 21."""

from __future__ import annotations

from datetime import datetime
from typing import Any

from utils.logger import get_logger

logger = get_logger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False


def _couleur_openpyxl(hex_str: str) -> str:
    """Retourne la couleur hex sans le '#' pour openpyxl."""
    return (hex_str or "#1f6aa5").lstrip("#").upper()


class ExcelDossierSubvention:
    """Générateur Excel multi-onglets pour le dossier de subvention."""

    def __init__(
        self,
        periode: str,
        type_periode: str = "scolaire",
        organisateur: str = "",
        objet: str = "",
        montant_demande: float = 0.0,
        sections: dict | None = None,
    ):
        self._periode = periode
        self._type_periode = type_periode
        self._organisateur = organisateur
        self._objet = objet
        self._montant_demande = float(montant_demande or 0)
        self._sections = sections or {}
        self._date_generation = datetime.now()

        from core.theme_export import get_theme_export
        theme = get_theme_export()
        self._couleur_principale = theme.get("couleur_principale", "#1f6aa5")
        self._couleur_secondaire = theme.get("couleur_secondaire", "#144870")

        self._date_debut, self._date_fin = self._calcul_periode()

    def _calcul_periode(self) -> tuple[str, str]:
        """Calcule les dates de début/fin selon le type de période."""
        if self._type_periode == "civile":
            try:
                annee = int(self._periode[:4])
                return f"{annee}-01-01", f"{annee}-12-31"
            except (ValueError, TypeError):
                pass
        else:
            # Année scolaire ex "2025-2026"
            try:
                parts = str(self._periode).split("-")
                annee_debut = int(parts[0])
                annee_fin = int(parts[1]) if len(parts) > 1 else annee_debut + 1
                return f"{annee_debut}-09-01", f"{annee_fin}-08-31"
            except (ValueError, IndexError):
                pass
        annee = datetime.now().year
        return f"{annee}-01-01", f"{annee}-12-31"

    def generer(self, chemin_sortie: str) -> bool:
        """Génère le fichier Excel et le sauvegarde.

        Returns:
            True si succès, False sinon.
        """
        if not _OPENPYXL_OK:
            logger.error("openpyxl non disponible")
            return False
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Supprime la feuille par défaut

            self._onglet_resume(wb)
            self._onglet_tresorerie(wb)
            self._onglet_par_categorie(wb)
            self._onglet_evenements(wb)
            self._onglet_subventions(wb)
            self._onglet_dons(wb)
            self._onglet_adherents(wb)
            self._onglet_budget_projet(wb)

            wb.save(chemin_sortie)
            logger.info("Excel dossier subvention sauvegardé : %s", chemin_sortie)
            return True
        except Exception as exc:
            logger.error("ExcelDossierSubvention.generer: %s", exc)
            return False

    # ── Styles communs ────────────────────────────────────────────────────────

    def _style_entete(self) -> tuple[Font, PatternFill, Alignment]:
        couleur = _couleur_openpyxl(self._couleur_principale)
        font = Font(bold=True, color="FFFFFF", size=11)
        fill = PatternFill("solid", fgColor=couleur)
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return font, fill, align

    def _style_total(self) -> tuple[Font, PatternFill]:
        couleur = _couleur_openpyxl(self._couleur_secondaire)
        font = Font(bold=True, color="FFFFFF", size=10)
        fill = PatternFill("solid", fgColor=couleur)
        return font, fill

    def _bordure_fine(self) -> Border:
        cote = Side(style="thin", color="CCCCCC")
        return Border(left=cote, right=cote, top=cote, bottom=cote)

    def _appliquer_entetes(self, ws, entetes: list[str], row: int = 1) -> None:
        font, fill, align = self._style_entete()
        border = self._bordure_fine()
        for col, texte in enumerate(entetes, 1):
            cell = ws.cell(row=row, column=col, value=texte)
            cell.font = font
            cell.fill = fill
            cell.alignment = align
            cell.border = border

    def _appliquer_ligne(self, ws, valeurs: list, row: int, gras: bool = False) -> None:
        border = self._bordure_fine()
        for col, val in enumerate(valeurs, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            if gras:
                cell.font = Font(bold=True)
            if isinstance(val, float):
                cell.number_format = '#,##0.00'

    def _auto_largeur(self, ws) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 50)

    # ── Onglets ───────────────────────────────────────────────────────────────

    def _onglet_resume(self, wb) -> None:
        ws = wb.create_sheet("Résumé")
        couleur = _couleur_openpyxl(self._couleur_principale)
        font_titre = Font(bold=True, color=couleur, size=14)
        align_centre = Alignment(horizontal="center")

        ws["A1"] = "DOSSIER DE DEMANDE DE SUBVENTION"
        ws["A1"].font = font_titre
        ws.merge_cells("A1:D1")
        ws["A1"].alignment = align_centre

        ws["A2"] = f"Organisme : {self._organisateur}"
        ws["A3"] = f"Objet : {self._objet}"
        ws["A4"] = f"Montant demandé : {self._montant_demande:,.2f} €"
        ws["A5"] = f"Période : {self._periode}"
        ws["A6"] = f"Généré le : {self._date_generation.strftime('%d/%m/%Y %H:%M')}"

        ws.append([])

        # KPI
        kpi_entetes = ["Indicateur", "Valeur"]
        self._appliquer_entetes(ws, kpi_entetes, ws.max_row + 1)

        kpis = self._charger_kpis()
        for lib, val in kpis:
            row_idx = ws.max_row + 1
            self._appliquer_ligne(ws, [lib, val], row_idx)

        self._auto_largeur(ws)

    def _charger_kpis(self) -> list[tuple[str, Any]]:
        """Charge les KPI depuis la base de données."""
        kpis = []
        try:
            from db.connection import get_connection
            conn = get_connection()
            try:
                # Adhérents
                row = conn.execute(
                    "SELECT COUNT(*) as nb FROM membres WHERE statut = 'actif'"
                ).fetchone()
                kpis.append(("Nombre d'adhérents actifs", row["nb"] if row else 0))

                # Événements sur la période
                row = conn.execute(
                    "SELECT COUNT(*) as nb FROM evenements WHERE date_debut >= ? AND date_debut <= ?",
                    (self._date_debut, self._date_fin),
                ).fetchone()
                kpis.append(("Nombre d'événements", row["nb"] if row else 0))

                # Recettes
                row = conn.execute(
                    """SELECT COALESCE(SUM(montant), 0) as total FROM operations_tresorerie
                       WHERE type = 'recette' AND date >= ? AND date <= ?""",
                    (self._date_debut, self._date_fin),
                ).fetchone()
                recettes = float(row["total"]) if row else 0.0
                kpis.append(("Total recettes", recettes))

                # Dépenses
                row = conn.execute(
                    """SELECT COALESCE(SUM(montant), 0) as total FROM operations_tresorerie
                       WHERE type = 'depense' AND date >= ? AND date <= ?""",
                    (self._date_debut, self._date_fin),
                ).fetchone()
                depenses = float(row["total"]) if row else 0.0
                kpis.append(("Total dépenses", depenses))
                kpis.append(("Solde", recettes - depenses))

                # Dons
                row = conn.execute(
                    """SELECT COALESCE(SUM(montant), 0) as total FROM dons
                       WHERE date_don >= ? AND date_don <= ?""",
                    (self._date_debut, self._date_fin),
                ).fetchone()
                kpis.append(("Total dons reçus", float(row["total"]) if row else 0.0))

            finally:
                conn.close()
        except Exception as exc:
            logger.error("_charger_kpis: %s", exc)
        return kpis

    def _onglet_tresorerie(self, wb) -> None:
        ws = wb.create_sheet("Trésorerie")
        entetes = ["Date", "Libellé", "Catégorie", "Type", "Montant (€)", "Compte"]
        self._appliquer_entetes(ws, entetes)

        try:
            from db.connection import get_connection
            conn = get_connection()
            try:
                rows = conn.execute(
                    """SELECT o.date, o.libelle, c.nom as categorie, o.type, o.montant,
                              cp.nom as compte
                       FROM operations_tresorerie o
                       LEFT JOIN categories_tresorerie c ON o.categorie_id = c.id
                       LEFT JOIN comptes cp ON o.compte_id = cp.id
                       WHERE o.date >= ? AND o.date <= ?
                       ORDER BY o.date""",
                    (self._date_debut, self._date_fin),
                ).fetchall()
            finally:
                conn.close()

            total_rec = 0.0
            total_dep = 0.0
            for row in rows:
                montant = float(row["montant"] or 0)
                row_idx = ws.max_row + 1
                self._appliquer_ligne(ws, [
                    row["date"],
                    row["libelle"],
                    row["categorie"] or "",
                    "Recette" if row["type"] == "recette" else "Dépense",
                    montant,
                    row["compte"] or "",
                ], row_idx)
                if row["type"] == "recette":
                    total_rec += montant
                else:
                    total_dep += montant

            # Ligne total
            row_idx = ws.max_row + 1
            font_t, fill_t = self._style_total()
            cells_tot = [
                ws.cell(row=row_idx, column=1, value="TOTAL"),
                ws.cell(row=row_idx, column=4, value="Recettes"),
                ws.cell(row=row_idx, column=5, value=total_rec),
            ]
            for c in cells_tot:
                c.font = font_t
                c.fill = fill_t

            row_idx += 1
            cells_tot2 = [
                ws.cell(row=row_idx, column=4, value="Dépenses"),
                ws.cell(row=row_idx, column=5, value=total_dep),
            ]
            for c in cells_tot2:
                c.font = font_t
                c.fill = fill_t

            row_idx += 1
            cells_tot3 = [
                ws.cell(row=row_idx, column=4, value="Solde"),
                ws.cell(row=row_idx, column=5, value=total_rec - total_dep),
            ]
            for c in cells_tot3:
                c.font = font_t
                c.fill = fill_t

        except Exception as exc:
            logger.error("_onglet_tresorerie: %s", exc)

        self._auto_largeur(ws)

    def _onglet_par_categorie(self, wb) -> None:
        ws = wb.create_sheet("Par catégorie")
        entetes = ["Catégorie", "Type", "Nombre d'opérations", "Total (€)"]
        self._appliquer_entetes(ws, entetes)

        try:
            from db.connection import get_connection
            conn = get_connection()
            try:
                rows = conn.execute(
                    """SELECT c.nom as categorie, o.type,
                              COUNT(*) as nb,
                              COALESCE(SUM(o.montant), 0) as total
                       FROM operations_tresorerie o
                       LEFT JOIN categories_tresorerie c ON o.categorie_id = c.id
                       WHERE o.date >= ? AND o.date <= ?
                       GROUP BY c.nom, o.type
                       ORDER BY c.nom, o.type""",
                    (self._date_debut, self._date_fin),
                ).fetchall()
            finally:
                conn.close()

            for row in rows:
                row_idx = ws.max_row + 1
                self._appliquer_ligne(ws, [
                    row["categorie"] or "(Sans catégorie)",
                    "Recette" if row["type"] == "recette" else "Dépense",
                    row["nb"],
                    float(row["total"]),
                ], row_idx)

        except Exception as exc:
            logger.error("_onglet_par_categorie: %s", exc)

        self._auto_largeur(ws)

    def _onglet_evenements(self, wb) -> None:
        ws = wb.create_sheet("Événements")
        entetes = ["Nom", "Date début", "Date fin", "Lieu", "Statut", "Participants"]
        self._appliquer_entetes(ws, entetes)

        try:
            from db.connection import get_connection
            conn = get_connection()
            try:
                rows = conn.execute(
                    """SELECT nom, date_debut, date_fin, lieu, statut,
                              nb_participants_total
                       FROM evenements
                       WHERE date_debut >= ? AND date_debut <= ?
                       ORDER BY date_debut""",
                    (self._date_debut, self._date_fin),
                ).fetchall()
            finally:
                conn.close()

            for row in rows:
                row_idx = ws.max_row + 1
                self._appliquer_ligne(ws, [
                    row["nom"],
                    row["date_debut"],
                    row["date_fin"] or "",
                    row["lieu"] or "",
                    row["statut"] or "",
                    row["nb_participants_total"] or 0,
                ], row_idx)

        except Exception as exc:
            logger.error("_onglet_evenements: %s", exc)

        self._auto_largeur(ws)

    def _onglet_subventions(self, wb) -> None:
        ws = wb.create_sheet("Subventions")
        entetes = ["Organisme", "Objet", "Montant demandé (€)", "Montant accordé (€)",
                   "Date demande", "Statut"]
        self._appliquer_entetes(ws, entetes)

        try:
            from db.connection import get_connection
            conn = get_connection()
            try:
                rows = conn.execute(
                    """SELECT organisme, objet, montant_demande, montant_accorde,
                              date_demande, statut
                       FROM subventions
                       WHERE date_demande >= ? AND date_demande <= ?
                       ORDER BY date_demande""",
                    (self._date_debut, self._date_fin),
                ).fetchall()
            finally:
                conn.close()

            for row in rows:
                row_idx = ws.max_row + 1
                self._appliquer_ligne(ws, [
                    row["organisme"] or "",
                    row["objet"] or "",
                    float(row["montant_demande"] or 0),
                    float(row["montant_accorde"] or 0),
                    row["date_demande"] or "",
                    row["statut"] or "",
                ], row_idx)

        except Exception as exc:
            logger.error("_onglet_subventions: %s", exc)

        self._auto_largeur(ws)

    def _onglet_dons(self, wb) -> None:
        ws = wb.create_sheet("Dons")
        entetes = ["Date", "Donateur", "Montant (€)", "Type", "Objet"]
        self._appliquer_entetes(ws, entetes)

        try:
            from db.connection import get_connection
            conn = get_connection()
            try:
                rows = conn.execute(
                    """SELECT date_don, donateur_nom, montant, type_don, objet
                       FROM dons
                       WHERE date_don >= ? AND date_don <= ?
                       ORDER BY date_don""",
                    (self._date_debut, self._date_fin),
                ).fetchall()
            finally:
                conn.close()

            total = 0.0
            for row in rows:
                montant = float(row["montant"] or 0)
                row_idx = ws.max_row + 1
                self._appliquer_ligne(ws, [
                    row["date_don"],
                    row["donateur_nom"] or "",
                    montant,
                    row["type_don"] or "",
                    row["objet"] or "",
                ], row_idx)
                total += montant

            # Total
            row_idx = ws.max_row + 1
            font_t, fill_t = self._style_total()
            for col, val in [(1, "TOTAL"), (3, total)]:
                c = ws.cell(row=row_idx, column=col, value=val)
                c.font = font_t
                c.fill = fill_t

        except Exception as exc:
            logger.error("_onglet_dons: %s", exc)

        self._auto_largeur(ws)

    def _onglet_adherents(self, wb) -> None:
        ws = wb.create_sheet("Adhérents")
        entetes = ["Indicateur", "Valeur"]
        self._appliquer_entetes(ws, entetes)

        try:
            from db.connection import get_connection
            conn = get_connection()
            try:
                # Total actifs
                row = conn.execute(
                    "SELECT COUNT(*) as nb FROM membres WHERE statut = 'actif'"
                ).fetchone()
                self._appliquer_ligne(ws, ["Membres actifs", row["nb"] if row else 0], ws.max_row + 1)

                # Par statut
                rows = conn.execute(
                    "SELECT statut, COUNT(*) as nb FROM membres GROUP BY statut ORDER BY statut"
                ).fetchall()
                for r in rows:
                    self._appliquer_ligne(ws, [
                        f"Statut : {r['statut']}",
                        r["nb"],
                    ], ws.max_row + 1)

            finally:
                conn.close()

        except Exception as exc:
            logger.error("_onglet_adherents: %s", exc)

        self._auto_largeur(ws)

    def _onglet_budget_projet(self, wb) -> None:
        ws = wb.create_sheet("Budget projet")

        ws["A1"] = "BUDGET PRÉVISIONNEL DU PROJET"
        couleur = _couleur_openpyxl(self._couleur_principale)
        ws["A1"].font = Font(bold=True, color=couleur, size=13)
        ws.merge_cells("A1:D1")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.append([])

        # Dépenses
        entetes_dep = ["Poste de dépense", "Montant prévu (€)", "Montant réel (€)", "Commentaire"]
        self._appliquer_entetes(ws, entetes_dep, ws.max_row + 1)

        lignes_dep = [
            ("Personnel", 0.0, 0.0, ""),
            ("Matériel", 0.0, 0.0, ""),
            ("Communication", 0.0, 0.0, ""),
            ("Déplacements", 0.0, 0.0, ""),
        ]
        for ligne in lignes_dep:
            self._appliquer_ligne(ws, list(ligne), ws.max_row + 1)

        # Total dépenses
        row_idx = ws.max_row + 1
        font_t, fill_t = self._style_total()
        tot_cells = [
            ws.cell(row=row_idx, column=1, value="Total dépenses"),
            ws.cell(row=row_idx, column=2, value=0.0),
            ws.cell(row=row_idx, column=3, value=0.0),
        ]
        for c in tot_cells:
            c.font = font_t
            c.fill = fill_t

        ws.append([])

        # Recettes
        entetes_rec = ["Source de financement", "Montant (€)", "", ""]
        self._appliquer_entetes(ws, entetes_rec, ws.max_row + 1)

        lignes_rec = [
            (f"Subvention demandée ({self._organisateur})", self._montant_demande),
            ("Cotisations", 0.0),
            ("Recettes propres", 0.0),
            ("Autres subventions", 0.0),
        ]
        for lib, val in lignes_rec:
            row_idx = ws.max_row + 1
            self._appliquer_ligne(ws, [lib, val, "", ""], row_idx)

        row_idx = ws.max_row + 1
        tot_cells2 = [
            ws.cell(row=row_idx, column=1, value="Total recettes"),
            ws.cell(row=row_idx, column=2, value=self._montant_demande),
        ]
        for c in tot_cells2:
            c.font = font_t
            c.fill = fill_t

        self._auto_largeur(ws)
