"""Orchestrateur central des exports (PDF et Excel) pour les événements."""

from __future__ import annotations

from dataclasses import dataclass

from utils.logger import get_logger

logger = get_logger(__name__)


# ── Configuration association ─────────────────────────────────────────────────


@dataclass
class ConfigAsso:
    nom: str
    adresse: str
    telephone: str
    email: str
    logo_path: str  # chemin fichier ou ""


def get_config_asso() -> ConfigAsso:
    """Lit les paramètres association depuis la DB."""
    from db.models.evenements import get_parametre

    return ConfigAsso(
        nom=get_parametre("asso_nom") or "",
        adresse=get_parametre("asso_adresse") or "",
        telephone=get_parametre("asso_telephone") or "",
        email=get_parametre("asso_email") or "",
        logo_path=get_parametre("asso_logo_path") or "",
    )


# ── Exports PDF ───────────────────────────────────────────────────────────────


def export_bilan_evenement_pdf(evenement_id: int, chemin_sortie: str) -> bool:
    """Génère le PDF complet du bilan d'un événement.

    Returns:
        True si succès, False sinon.
    """
    try:
        from core.pdf_generator import PdfEvenement

        config = get_config_asso()
        gen = PdfEvenement(evenement_id, config)
        return gen.generer(chemin_sortie)
    except Exception as exc:
        logger.error("export_bilan_evenement_pdf: %s", exc)
        return False


def export_pv_tirage_pdf(evenement_id: int, chemin_sortie: str) -> bool:
    """Génère le PV de tirage tombola en PDF.

    Returns:
        True si succès, False sinon.
    """
    try:
        from core.pdf_generator import PvTirage

        config = get_config_asso()
        gen = PvTirage(evenement_id, config)
        return gen.generer(chemin_sortie)
    except Exception as exc:
        logger.error("export_pv_tirage_pdf: %s", exc)
        return False


def export_liste_benevoles_pdf(evenement_id: int, chemin_sortie: str) -> bool:
    """Génère la liste des bénévoles en PDF.

    Returns:
        True si succès, False sinon.
    """
    try:
        from core.pdf_generator import ListeBenevolesPdf

        config = get_config_asso()
        gen = ListeBenevolesPdf(evenement_id, config)
        return gen.generer(chemin_sortie)
    except Exception as exc:
        logger.error("export_liste_benevoles_pdf: %s", exc)
        return False


# ── Exports Excel ─────────────────────────────────────────────────────────────


def export_bilan_evenement_excel(evenement_id: int, chemin_sortie: str) -> bool:
    """Génère le classeur Excel complet d'un événement.

    Returns:
        True si succès, False sinon.
    """
    try:
        from core.excel_generator import ExcelEvenement

        config = get_config_asso()
        gen = ExcelEvenement(evenement_id, config)
        return gen.generer(chemin_sortie)
    except Exception as exc:
        logger.error("export_bilan_evenement_excel: %s", exc)
        return False


def export_liste_benevoles_excel(evenement_id: int, chemin_sortie: str) -> bool:
    """Génère la liste des bénévoles en Excel.

    Returns:
        True si succès, False sinon.
    """
    try:
        from core.excel_generator import ExcelBenevoles

        config = get_config_asso()
        gen = ExcelBenevoles(evenement_id, config)
        return gen.generer(chemin_sortie)
    except Exception as exc:
        logger.error("export_liste_benevoles_excel: %s", exc)
        return False


# ── Utilitaires ───────────────────────────────────────────────────────────────


def slugifier_nom(nom: str) -> str:
    """Convertit un nom en slug sûr pour les noms de fichiers.

    Minuscules, espaces → underscores, suppression des accents et caractères spéciaux.
    """
    import re
    import unicodedata

    # Normalisation unicode → suppression accents
    nfd = unicodedata.normalize("NFD", nom)
    sans_accents = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
    # Minuscules
    slug = sans_accents.lower()
    # Espaces → underscores
    slug = slug.replace(" ", "_")
    # Suppression des caractères non alphanumériques (sauf underscore)
    slug = re.sub(r"[^a-z0-9_]", "", slug)
    # Suppression des underscores multiples
    slug = re.sub(r"_+", "_", slug)
    slug = slug.strip("_")
    return slug or "evenement"


def generer_nom_fichier(nom_evenement: str, date: str, extension: str) -> str:
    """Génère un nom de fichier normalisé pour l'export.

    Format : bilan_{slug}_{date}.{extension}
    """
    slug = slugifier_nom(nom_evenement)
    return f"bilan_{slug}_{date}.{extension}"


def montant_signe_operation(operation: dict) -> float:
    """Retourne le montant signé d'une opération de trésorerie.

    Recettes : positif. Dépenses et virements sortants : négatif.
    Opérations non validées : 0.
    """
    montant = float(operation.get("montant") or 0)
    if operation.get("statut") != "valide":
        return 0.0
    type_operation = operation.get("type_operation")
    if type_operation == "recette":
        return montant
    if type_operation == "depense":
        return -montant
    if type_operation == "virement_interne":
        return montant if operation.get("source_module") == "virement_entrant" else -montant
    return montant



def export_bilan_ag_pdf(exercice: str, chemin_sortie: str,
                        sections: dict | None = None,
                        avec_graphiques: bool = False,
                        type_periode: str = "scolaire") -> bool:
    """Génère le bilan AG au format PDF."""
    try:
        from core.pdf_bilan_ag import PdfBilanAG

        gen = PdfBilanAG(exercice, sections=sections, avec_graphiques=avec_graphiques,
                         type_periode=type_periode)
        return gen.generer(chemin_sortie)
    except Exception as exc:
        logger.exception("export_bilan_ag_pdf: %s", exc)
        return False


def export_liste_adherents_pdf(chemin_sortie: str, filtre_statut: str = '') -> bool:
    """Génère la liste des adhérents en PDF."""
    try:
        from core.pdf_adherents import PdfListeAdherents

        return PdfListeAdherents(filtre_statut=filtre_statut).generer(chemin_sortie)
    except Exception as exc:
        logger.error("export_liste_adherents_pdf: %s", exc)
        return False


def export_fiche_adherent_pdf(membre_id: int, chemin_sortie: str) -> bool:
    """Génère la fiche d'un adhérent en PDF."""
    try:
        from core.pdf_adherents import PdfFicheAdherent

        return PdfFicheAdherent(membre_id).generer(chemin_sortie)
    except Exception as exc:
        logger.error("export_fiche_adherent_pdf: %s", exc)
        return False


def export_liste_stock_pdf(chemin_sortie: str) -> bool:
    """Génère la liste du stock en PDF."""
    try:
        from core.pdf_stock import PdfListeStock

        return PdfListeStock().generer(chemin_sortie)
    except Exception as exc:
        logger.error("export_liste_stock_pdf: %s", exc)
        return False


def export_historique_stock_pdf(chemin_sortie: str, date_debut: str = '', date_fin: str = '') -> bool:
    """Génère l'historique du stock en PDF."""
    try:
        from core.pdf_stock import PdfHistoriqueStock

        return PdfHistoriqueStock(date_debut=date_debut, date_fin=date_fin).generer(chemin_sortie)
    except Exception as exc:
        logger.error("export_historique_stock_pdf: %s", exc)
        return False


def export_releve_compte_pdf(compte_id: int, chemin_sortie: str,
                              date_debut: str = '', date_fin: str = '') -> bool:
    """Génère le relevé PDF d'un compte."""
    try:
        from core.pdf_tresorerie import PdfReleverCompte

        return PdfReleverCompte(compte_id, date_debut=date_debut, date_fin=date_fin).generer(chemin_sortie)
    except Exception as exc:
        logger.error("export_releve_compte_pdf: %s", exc)
        return False


def export_subventions_pdf(chemin_sortie: str, exercice: str = '') -> bool:
    """Génère le récapitulatif PDF des subventions."""
    try:
        from core.pdf_tresorerie import PdfSubventions

        return PdfSubventions(exercice=exercice).generer(chemin_sortie)
    except Exception as exc:
        logger.error("export_subventions_pdf: %s", exc)
        return False


def export_rapport_caisse_pdf(chemin_sortie: str, session_id: int | None = None) -> bool:
    """Génère le rapport PDF de caisse buvette."""
    try:
        from core.pdf_buvette import PdfRapportCaisse

        return PdfRapportCaisse(session_id=session_id).generer(chemin_sortie)
    except Exception as exc:
        logger.error("export_rapport_caisse_pdf: %s", exc)
        return False


def _exporter_classeur_phase9(
    chemin_sortie: str,
    titre: str,
    colonnes: list[str],
    lignes: list[list],
    feuille: str = 'Export',
    colonnes_montant: set[int] | None = None,
) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = feuille[:31] or 'Export'
        ws.append([titre])
        ws['A1'].font = Font(bold=True, size=14, color='000000')
        ws.append([])
        ws.append(colonnes)

        fill_entete = PatternFill('solid', fgColor='F0F0F0')
        fill_total = PatternFill('solid', fgColor='E0E0E0')
        font_entete = Font(bold=True, color='000000')
        font_bold = Font(bold=True, color='000000')
        align_top = Alignment(vertical='top')
        align_right = Alignment(horizontal='right', vertical='top')
        colonnes_montant = colonnes_montant or set()

        for cell in ws[3]:
            cell.fill = fill_entete
            cell.font = font_entete
            cell.alignment = align_top

        for ligne in lignes:
            ws.append(ligne)

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            est_total = str(row[0].value or '').strip().lower() == 'total'
            for idx, cell in enumerate(row, start=1):
                cell.alignment = align_right if idx in colonnes_montant else align_top
                if idx in colonnes_montant and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00 €'
                if est_total:
                    cell.fill = fill_total
                    cell.font = font_bold

        for column_cells in ws.columns:
            col_letter = get_column_letter(column_cells[0].column)
            max_len = max(len(str(cell.value or '')) for cell in column_cells)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)

        wb.save(chemin_sortie)
        return True
    except Exception as exc:
        logger.error("_exporter_classeur_phase9: %s", exc)
        return False


def export_liste_adherents_excel(chemin_sortie: str) -> bool:
    """Génère la liste des adhérents en Excel."""
    try:
        from db.models.membres import get_all_membres

        membres = get_all_membres(include_archives=False)
        lignes = [
            [
                membre.get('nom') or '',
                membre.get('prenom') or '',
                membre.get('statut') or '',
                membre.get('telephone') or '',
                membre.get('email') or '',
            ]
            for membre in membres
        ]
        lignes.append(['Total', '', len(membres), '', ''])
        return _exporter_classeur_phase9(
            chemin_sortie,
            'Liste des adhérents',
            ['Nom', 'Prénom', 'Statut', 'Téléphone', 'Email'],
            lignes,
            feuille='Adherents',
        )
    except Exception as exc:
        logger.error("export_liste_adherents_excel: %s", exc)
        return False


def export_tresorerie_excel(chemin_sortie: str, compte_id: int | None = None,
                             date_debut: str = '', date_fin: str = '') -> bool:
    """Génère l'export Excel des opérations de trésorerie."""
    try:
        from db.models.tresorerie import get_operations

        operations = get_operations(
            compte_id=compte_id,
            date_debut=date_debut or None,
            date_fin=date_fin or None,
        )
        lignes = []
        total = 0.0
        for operation in operations:
            montant = montant_signe_operation(operation)
            total += montant
            lignes.append([
                operation.get('date_operation') or '',
                operation.get('compte_nom') or '',
                operation.get('libelle') or '',
                operation.get('categorie_nom') or '',
                operation.get('type_operation') or '',
                montant,
                operation.get('statut') or '',
            ])
        lignes.append(['Total', '', '', '', '', total, ''])
        return _exporter_classeur_phase9(
            chemin_sortie,
            'Trésorerie',
            ['Date', 'Compte', 'Libellé', 'Catégorie', 'Type', 'Montant', 'Statut'],
            lignes,
            feuille='Tresorerie',
            colonnes_montant={6},
        )
    except Exception as exc:
        logger.error("export_tresorerie_excel: %s", exc)
        return False


def export_stock_excel(chemin_sortie: str) -> bool:
    """Génère l'export Excel du stock."""
    try:
        from db.models.stock import get_all_articles

        articles = get_all_articles(include_archives=False)
        lignes = []
        for article in articles:
            quantite = int(article.get('quantite') or 0)
            seuil = int(article.get('seuil_alerte') or 0)
            statut = 'Sous seuil' if quantite <= seuil else 'OK'
            lignes.append([
                article.get('nom') or '',
                article.get('categorie_nom') or '',
                quantite,
                article.get('unite_nom') or '',
                seuil,
                statut,
            ])
        lignes.append(['Total', '', len(articles), '', '', ''])
        return _exporter_classeur_phase9(
            chemin_sortie,
            'Liste du stock',
            ['Désignation', 'Catégorie', 'Qté', 'Unité', 'Seuil', 'Statut'],
            lignes,
            feuille='Stock',
        )
    except Exception as exc:
        logger.error("export_stock_excel: %s", exc)
        return False


def generer_nom_fichier_phase9(prefixe: str, suffixe: str = '', extension: str = 'pdf') -> str:
    """Génère un nom de fichier normalisé: prefixe_suffixe_YYYY-MM-DD.extension"""
    from datetime import datetime

    morceaux = [slugifier_nom(prefixe)]
    if suffixe.strip():
        morceaux.append(slugifier_nom(suffixe))
    return f"{'_'.join(morceaux)}_{datetime.now().strftime('%Y-%m-%d')}.{extension}"
