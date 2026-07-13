"""Générateurs Excel pour les exports d'événements.

Utilise openpyxl (Workbook, PatternFill, Font, Alignment).
Aucun import tkinter/customtkinter.
"""

from __future__ import annotations

import re
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.exports import montant_signe_operation
from utils.logger import get_logger

if TYPE_CHECKING:
    from core.exports import ConfigAsso

logger = get_logger(__name__)

# ── Constantes de style ───────────────────────────────────────────────────────

COULEUR_ENTETE = "1f6aa5"
COULEUR_TOTAL = "d0e4f7"
COULEUR_SOUS_TITRE = "dce9f5"

_FILL_ENTETE = PatternFill("solid", fgColor=COULEUR_ENTETE)
_FILL_TOTAL = PatternFill("solid", fgColor=COULEUR_TOTAL)
_FILL_SOUS_TITRE = PatternFill("solid", fgColor=COULEUR_SOUS_TITRE)

_FONT_ENTETE = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_FONT_BOLD = Font(name="Calibri", bold=True, size=11)
_FONT_NORMAL = Font(name="Calibri", size=11)
_FONT_TITRE = Font(name="Calibri", bold=True, size=13, color=COULEUR_ENTETE)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def _formater_montant(v) -> float | str:
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return ""


def _formater_date(v: str | None) -> str:
    if not v:
        return ""
    from datetime import datetime

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(v, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return str(v)


def _ecrire_entete_colonne(ws, row: int, colonnes: list[str]) -> None:
    """Écrit une ligne d'en-tête avec style."""
    for col_idx, nom in enumerate(colonnes, start=1):
        cell = ws.cell(row=row, column=col_idx, value=nom)
        cell.font = _FONT_ENTETE
        cell.fill = _FILL_ENTETE
        cell.alignment = _ALIGN_CENTER


def _ajuster_largeurs(ws) -> None:
    """Ajuste automatiquement la largeur des colonnes."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 60)


def _ligne_total(ws, row: int, nb_cols: int, label: str, valeur_montant: float | None = None) -> None:
    """Écrit une ligne de total en gras."""
    cell_label = ws.cell(row=row, column=1, value=label)
    cell_label.font = _FONT_BOLD
    cell_label.fill = _FILL_TOTAL
    if valeur_montant is not None and nb_cols >= 2:
        cell_val = ws.cell(row=row, column=nb_cols, value=valeur_montant)
        cell_val.font = _FONT_BOLD
        cell_val.fill = _FILL_TOTAL
        cell_val.number_format = '#,##0.00 €'


# ── ExcelEvenement ────────────────────────────────────────────────────────────


class ExcelEvenement:
    """Génère le classeur Excel complet d'un événement."""

    def __init__(self, evenement_id: int, config_asso: "ConfigAsso") -> None:
        self._evenement_id = evenement_id
        self._config = config_asso

    def generer(self, chemin_sortie: str) -> bool:
        """Génère le classeur Excel et l'enregistre dans chemin_sortie."""
        try:
            from db.models.evenements import (
                get_evenement_by_id,
                get_depenses_evenement,
                get_benevoles_evenement,
            )
            from db.models.tombola import get_lots_evenement, get_carnets_evenement
            from db.models.stands import get_stands_evenement
            from db.models.tableaux import get_tableaux_evenement, get_colonnes_tableau, get_lignes_tableau
            from core.evenements import calculer_bilan_evenement

            evenement = get_evenement_by_id(self._evenement_id)
            if not evenement:
                logger.error("ExcelEvenement: événement %s introuvable", self._evenement_id)
                return False

            wb = Workbook()
            wb.remove(wb.active)  # Supprime la feuille vide par défaut

            # ── Onglet Résumé ──────────────────────────────────────────────────
            ws_resume = wb.create_sheet("Résumé")
            self._construire_resume(ws_resume, evenement, calculer_bilan_evenement(self._evenement_id))

            # ── Onglet Billetterie ─────────────────────────────────────────────
            from db.models.evenements import get_ventes_evenement, get_lignes_vente, get_stats_billetterie
            ws_bill = wb.create_sheet("Billetterie")
            ventes = get_ventes_evenement(self._evenement_id)
            stats_bill = get_stats_billetterie(self._evenement_id)
            self._construire_billetterie(ws_bill, ventes, stats_bill, get_lignes_vente)

            # ── Onglet Dépenses ────────────────────────────────────────────────
            ws_dep = wb.create_sheet("Dépenses")
            depenses = get_depenses_evenement(self._evenement_id)
            self._construire_depenses(ws_dep, depenses)

            # ── Onglet Bénévoles ───────────────────────────────────────────────
            ws_ben = wb.create_sheet("Bénévoles")
            benevoles = get_benevoles_evenement(self._evenement_id)
            self._construire_benevoles(ws_ben, benevoles)

            # ── Tombola ────────────────────────────────────────────────────────
            carnets = get_carnets_evenement(self._evenement_id)
            lots = get_lots_evenement(self._evenement_id)
            if carnets:
                ws_carn = wb.create_sheet("Tombola_Carnets")
                self._construire_tombola_carnets(ws_carn, carnets)
            if lots:
                ws_lots = wb.create_sheet("Tombola_Lots")
                self._construire_tombola_lots(ws_lots, lots)

            # ── Stands ────────────────────────────────────────────────────────
            stands = get_stands_evenement(self._evenement_id)
            if stands:
                ws_stands = wb.create_sheet("Stands")
                self._construire_stands(ws_stands, stands)

            # ── Tableaux personnalisés ─────────────────────────────────────────
            tableaux = get_tableaux_evenement(self._evenement_id)
            for tableau in tableaux:
                colonnes = get_colonnes_tableau(int(tableau["id"]))
                lignes = get_lignes_tableau(int(tableau["id"]))
                if not colonnes:
                    continue
                # Nom d'onglet : max 31 chars, caractères spéciaux interdits
                nom_onglet = str(tableau.get("nom") or "Tableau")[:31]
                nom_onglet = re.sub(r"[/\\?*\[\]:]", "", nom_onglet)
                ws_tab = wb.create_sheet(nom_onglet)
                self._construire_tableau_perso(ws_tab, tableau, colonnes, lignes)

            wb.save(chemin_sortie)
            return True
        except Exception as exc:
            logger.error("ExcelEvenement.generer: %s", exc)
            return False

    def _construire_resume(self, ws, evenement: dict, bilan: dict) -> None:
        """Construit l'onglet Résumé."""
        LABELS_STATUT = {
            "planifie": "Planifié",
            "en_cours": "En cours",
            "termine": "Terminé",
            "annule": "Annulé",
        }
        row = 1
        titre_cell = ws.cell(row=row, column=1, value=f"Bilan — {evenement.get('nom') or '—'}")
        titre_cell.font = _FONT_TITRE
        row += 2

        ws.cell(row=row, column=1, value="Informations générales").font = _FONT_BOLD
        row += 1
        infos = [
            ("Nom", evenement.get("nom") or "—"),
            ("Type", evenement.get("type") or "—"),
            ("Date début", _formater_date(evenement.get("date_debut"))),
            ("Date fin", _formater_date(evenement.get("date_fin"))),
            ("Statut", LABELS_STATUT.get(evenement.get("statut") or "", evenement.get("statut") or "—")),
            ("Budget prévisionnel", _formater_montant(evenement.get("budget_previsionnel"))),
        ]
        for label, val in infos:
            ws.cell(row=row, column=1, value=label).font = _FONT_BOLD
            ws.cell(row=row, column=2, value=val)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Résumé financier").font = _FONT_BOLD
        row += 1
        _ecrire_entete_colonne(ws, row, ["", "Montant"])
        row += 1
        financiers = [
            ("Recettes totales", bilan["recettes_total"]),
            ("Dépenses totales", bilan["depenses_total"]),
            ("Bénéfice net", bilan["benefice"]),
        ]
        for i, (label, val) in enumerate(financiers):
            ws.cell(row=row, column=1, value=label)
            cell_val = ws.cell(row=row, column=2, value=val)
            cell_val.number_format = '#,##0.00 €'
            if i == len(financiers) - 1:
                ws.cell(row=row, column=1).font = _FONT_BOLD
                ws.cell(row=row, column=2).font = _FONT_BOLD
                ws.cell(row=row, column=1).fill = _FILL_TOTAL
                ws.cell(row=row, column=2).fill = _FILL_TOTAL
            row += 1

        _ajuster_largeurs(ws)

    def _construire_billetterie(self, ws, ventes: list, stats: dict, get_lignes_vente_fn) -> None:
        """Construit l'onglet Billetterie."""
        CANAUX = {"sur_place": "Sur place", "prevente": "Prévente"}
        MODES = {"especes": "Espèces", "cheque": "Chèque", "carte": "Carte", "sumup": "SumUp"}
        row = 1
        ws.cell(row=row, column=1, value="Ventes").font = _FONT_BOLD
        row += 1
        headers = ["Date", "Canal", "Mode paiement", "Tarif", "Quantité", "Prix unit.", "Sous-total", "Statut"]
        _ecrire_entete_colonne(ws, row, headers)
        row += 1
        for vente in ventes:
            lignes = get_lignes_vente_fn(int(vente["id"]))
            date_str = _formater_date(vente.get("date"))
            canal = CANAUX.get(vente.get("canal") or "", vente.get("canal") or "—")
            mode = MODES.get(vente.get("mode_paiement") or "", vente.get("mode_paiement") or "—")
            statut = vente.get("statut") or "—"
            for ligne in lignes:
                ws.cell(row=row, column=1, value=date_str)
                ws.cell(row=row, column=2, value=canal)
                ws.cell(row=row, column=3, value=mode)
                ws.cell(row=row, column=4, value=ligne.get("tarif_nom") or "—")
                ws.cell(row=row, column=5, value=int(ligne.get("quantite") or 0))
                cell_pu = ws.cell(row=row, column=6, value=_formater_montant(ligne.get("prix_unitaire")))
                cell_pu.number_format = '#,##0.00 €'
                cell_st = ws.cell(row=row, column=7, value=_formater_montant(ligne.get("sous_total")))
                cell_st.number_format = '#,##0.00 €'
                ws.cell(row=row, column=8, value=statut)
                row += 1
        # Totaux
        total_net = _formater_montant(stats.get("total_net"))
        if isinstance(total_net, float):
            row_total = row
            ws.cell(row=row_total, column=1, value=f"Total net : {total_net:,.2f} €").font = _FONT_BOLD
            ws.cell(row=row_total, column=1).fill = _FILL_TOTAL
        _ajuster_largeurs(ws)

    def _construire_depenses(self, ws, depenses: list) -> None:
        """Construit l'onglet Dépenses."""
        row = 1
        headers = ["Libellé", "Montant", "Fournisseur", "Mode paiement", "Date", "Catégorie", "Commentaire"]
        _ecrire_entete_colonne(ws, row, headers)
        row += 1
        total = 0.0
        for d in depenses:
            ws.cell(row=row, column=1, value=d.get("libelle") or "—")
            montant = _formater_montant(d.get("montant"))
            cell_m = ws.cell(row=row, column=2, value=montant)
            cell_m.number_format = '#,##0.00 €'
            if isinstance(montant, float):
                total += montant
            ws.cell(row=row, column=3, value=d.get("fournisseur_nom") or "—")
            ws.cell(row=row, column=4, value=d.get("mode_paiement") or "—")
            ws.cell(row=row, column=5, value=_formater_date(d.get("date")))
            ws.cell(row=row, column=6, value=d.get("categorie") or "—")
            ws.cell(row=row, column=7, value=d.get("commentaire") or "")
            row += 1
        # Total
        ws.cell(row=row, column=1, value="Total").font = _FONT_BOLD
        ws.cell(row=row, column=1).fill = _FILL_TOTAL
        cell_tot = ws.cell(row=row, column=2, value=round(total, 2))
        cell_tot.font = _FONT_BOLD
        cell_tot.fill = _FILL_TOTAL
        cell_tot.number_format = '#,##0.00 €'
        _ajuster_largeurs(ws)

    def _construire_benevoles(self, ws, benevoles: list) -> None:
        """Construit l'onglet Bénévoles."""
        STATUTS_BEN = {"confirme": "Confirmé", "desiste": "Désisté", "remplace": "Remplacé"}
        row = 1
        headers = ["Nom", "Prénom", "Rôle", "Heure début", "Heure fin", "Statut", "Commentaire"]
        _ecrire_entete_colonne(ws, row, headers)
        row += 1
        for b in benevoles:
            if b.get("membre_id"):
                nom = b.get("membre_nom") or ""
                prenom = b.get("membre_prenom") or ""
            else:
                nom = b.get("nom_externe") or ""
                prenom = b.get("prenom_externe") or ""
            ws.cell(row=row, column=1, value=nom)
            ws.cell(row=row, column=2, value=prenom)
            ws.cell(row=row, column=3, value=b.get("role") or "—")
            ws.cell(row=row, column=4, value=b.get("heure_debut") or "")
            ws.cell(row=row, column=5, value=b.get("heure_fin") or "")
            ws.cell(row=row, column=6, value=STATUTS_BEN.get(b.get("statut") or "", b.get("statut") or "—"))
            ws.cell(row=row, column=7, value=b.get("commentaire") or "")
            row += 1
        # Total
        ws.cell(row=row, column=1, value=f"Total : {len(benevoles)} bénévole(s)").font = _FONT_BOLD
        ws.cell(row=row, column=1).fill = _FILL_TOTAL
        _ajuster_largeurs(ws)

    def _construire_tombola_carnets(self, ws, carnets: list) -> None:
        """Construit l'onglet Tombola_Carnets."""
        STATUTS_CARNET = {"emis": "Émis", "vendu": "Vendu", "retourne": "Retourné", "perdu": "Perdu"}
        row = 1
        headers = ["N° début", "N° fin", "Prix carnet", "Vendeur", "Statut", "Encaissé", "Date remise"]
        _ecrire_entete_colonne(ws, row, headers)
        row += 1
        for c in carnets:
            vendeur = ""
            if c.get("vendeur_membre_id"):
                vendeur = f"{c.get('vendeur_prenom') or ''} {c.get('vendeur_nom') or ''}".strip()
            else:
                vendeur = c.get("vendeur_nom_externe") or ""
            ws.cell(row=row, column=1, value=int(c.get("numero_debut") or 0))
            ws.cell(row=row, column=2, value=int(c.get("numero_fin") or 0))
            cell_prix = ws.cell(row=row, column=3, value=_formater_montant(c.get("prix_carnet")))
            cell_prix.number_format = '#,##0.00 €'
            ws.cell(row=row, column=4, value=vendeur or "—")
            ws.cell(row=row, column=5, value=STATUTS_CARNET.get(c.get("statut") or "", c.get("statut") or "—"))
            cell_enc = ws.cell(row=row, column=6, value=_formater_montant(c.get("montant_encaisse")))
            cell_enc.number_format = '#,##0.00 €'
            ws.cell(row=row, column=7, value=_formater_date(c.get("date_remise")))
            row += 1
        _ajuster_largeurs(ws)

    def _construire_tombola_lots(self, ws, lots: list) -> None:
        """Construit l'onglet Tombola_Lots."""
        STATUTS_LOT = {"en_attente": "En attente", "attribue": "Attribué", "non_reclame": "Non réclamé"}
        TYPES_LOT = {"achete": "Acheté", "sponsorise": "Sponsorisé"}
        row = 1
        headers = ["N°", "Description", "Valeur estimée", "Type", "Sponsor/Fourni", "N° gagnant", "Statut", "Date tirage"]
        _ecrire_entete_colonne(ws, row, headers)
        row += 1
        for lot in lots:
            ws.cell(row=row, column=1, value=int(lot.get("numero") or 0))
            ws.cell(row=row, column=2, value=lot.get("description") or "—")
            cell_val = ws.cell(row=row, column=3, value=_formater_montant(lot.get("valeur_estimee")))
            cell_val.number_format = '#,##0.00 €'
            ws.cell(row=row, column=4, value=TYPES_LOT.get(lot.get("type_lot") or "", lot.get("type_lot") or "—"))
            ws.cell(row=row, column=5, value=lot.get("sponsor_nom") or lot.get("fournisseur_nom") or "—")
            ws.cell(row=row, column=6, value=lot.get("numero_gagnant") or "—")
            ws.cell(row=row, column=7, value=STATUTS_LOT.get(lot.get("statut") or "", lot.get("statut") or "—"))
            ws.cell(row=row, column=8, value=_formater_date(lot.get("date_tirage")))
            row += 1
        _ajuster_largeurs(ws)

    def _construire_stands(self, ws, stands: list) -> None:
        """Construit l'onglet Stands."""
        TYPES_STAND = {"benevole": "Bénévole", "location": "Location"}
        STATUTS_STAND = {"confirme": "Confirmé", "annule": "Annulé"}
        row = 1
        headers = ["N° emplacement", "Nom stand", "Type", "Responsable", "Montant location", "Paiement avant", "Statut"]
        _ecrire_entete_colonne(ws, row, headers)
        row += 1
        total_locations = 0.0
        for s in stands:
            if s.get("responsable_membre_id"):
                resp = f"{s.get('responsable_prenom') or ''} {s.get('responsable_nom') or ''}".strip()
            else:
                resp = s.get("responsable_nom_externe") or "—"
            ws.cell(row=row, column=1, value=s.get("numero_emplacement") or "—")
            ws.cell(row=row, column=2, value=s.get("nom_stand") or "—")
            ws.cell(row=row, column=3, value=TYPES_STAND.get(s.get("type_stand") or "", s.get("type_stand") or "—"))
            ws.cell(row=row, column=4, value=resp or "—")
            montant = _formater_montant(s.get("montant_location")) if s.get("type_stand") == "location" else ""
            if isinstance(montant, float):
                total_locations += montant
                cell_m = ws.cell(row=row, column=5, value=montant)
                cell_m.number_format = '#,##0.00 €'
            else:
                ws.cell(row=row, column=5, value="—")
            ws.cell(row=row, column=6, value="Oui" if s.get("paiement_avant") else "Non")
            ws.cell(row=row, column=7, value=STATUTS_STAND.get(s.get("statut") or "", s.get("statut") or "—"))
            row += 1
        # Total locations
        ws.cell(row=row, column=1, value="Total locations").font = _FONT_BOLD
        ws.cell(row=row, column=1).fill = _FILL_TOTAL
        cell_tot = ws.cell(row=row, column=5, value=round(total_locations, 2))
        cell_tot.font = _FONT_BOLD
        cell_tot.fill = _FILL_TOTAL
        cell_tot.number_format = '#,##0.00 €'
        _ajuster_largeurs(ws)

    def _construire_tableau_perso(self, ws, tableau: dict, colonnes: list, lignes: list) -> None:
        """Construit un onglet pour un tableau personnalisé."""
        row = 1
        nom_tab = tableau.get("nom") or "Tableau"
        ws.cell(row=row, column=1, value=nom_tab).font = _FONT_TITRE
        row += 1
        if tableau.get("description"):
            ws.cell(row=row, column=1, value=tableau["description"])
            row += 1
        row += 1

        headers = [c.get("nom") or "" for c in colonnes]
        _ecrire_entete_colonne(ws, row, headers)
        row += 1

        for ligne in lignes:
            cellules = ligne.get("cellules") or {}
            for col_idx, col in enumerate(colonnes, start=1):
                val = cellules.get(str(col["id"])) or ""
                cell = ws.cell(row=row, column=col_idx, value=val)
                if col.get("type_colonne") in ("montant", "nombre"):
                    try:
                        cell.value = float(val) if val else ""
                        if col.get("type_colonne") == "montant":
                            cell.number_format = '#,##0.00 €'
                    except (ValueError, TypeError):
                        pass
            row += 1

        # Totaux colonnes montant/nombre avec afficher_total
        cols_total = [c for c in colonnes if c.get("afficher_total") and c.get("type_colonne") in ("montant", "nombre")]
        if cols_total:
            for col in cols_total:
                col_idx = colonnes.index(col) + 1
                total = 0.0
                for ligne in lignes:
                    val = (ligne.get("cellules") or {}).get(str(col["id"])) or ""
                    try:
                        total += float(val)
                    except (ValueError, TypeError):
                        pass
                cell_tot = ws.cell(row=row, column=col_idx, value=round(total, 2))
                cell_tot.font = _FONT_BOLD
                cell_tot.fill = _FILL_TOTAL
                if col.get("type_colonne") == "montant":
                    cell_tot.number_format = '#,##0.00 €'
        if cols_total:
            ws.cell(row=row, column=1, value="Total").font = _FONT_BOLD
            ws.cell(row=row, column=1).fill = _FILL_TOTAL

        _ajuster_largeurs(ws)


# ── ExcelBenevoles ────────────────────────────────────────────────────────────


class ExcelBenevoles:
    """Génère la liste des bénévoles en Excel."""

    def __init__(self, evenement_id: int, config_asso: "ConfigAsso") -> None:
        self._evenement_id = evenement_id
        self._config = config_asso

    def generer(self, chemin_sortie: str) -> bool:
        """Génère la liste des bénévoles et l'enregistre dans chemin_sortie."""
        try:
            from db.models.evenements import get_evenement_by_id, get_benevoles_evenement

            evenement = get_evenement_by_id(self._evenement_id)
            nom_ev = (evenement or {}).get("nom") or f"Événement #{self._evenement_id}"
            benevoles = get_benevoles_evenement(self._evenement_id)

            wb = Workbook()
            ws = wb.active
            ws.title = "Bénévoles"

            row = 1
            ws.cell(row=row, column=1, value=f"Liste des Bénévoles — {nom_ev}").font = _FONT_TITRE
            row += 2

            STATUTS_BEN = {"confirme": "Confirmé", "desiste": "Désisté", "remplace": "Remplacé"}
            headers = ["Nom", "Prénom", "Rôle", "Heure début", "Heure fin", "Statut", "Commentaire"]
            _ecrire_entete_colonne(ws, row, headers)
            row += 1

            for b in benevoles:
                if b.get("membre_id"):
                    nom = b.get("membre_nom") or ""
                    prenom = b.get("membre_prenom") or ""
                else:
                    nom = b.get("nom_externe") or ""
                    prenom = b.get("prenom_externe") or ""
                ws.cell(row=row, column=1, value=nom)
                ws.cell(row=row, column=2, value=prenom)
                ws.cell(row=row, column=3, value=b.get("role") or "—")
                ws.cell(row=row, column=4, value=b.get("heure_debut") or "")
                ws.cell(row=row, column=5, value=b.get("heure_fin") or "")
                ws.cell(row=row, column=6, value=STATUTS_BEN.get(b.get("statut") or "", b.get("statut") or "—"))
                ws.cell(row=row, column=7, value=b.get("commentaire") or "")
                row += 1

            # Total
            ws.cell(row=row, column=1, value=f"Total : {len(benevoles)} bénévole(s)").font = _FONT_BOLD
            ws.cell(row=row, column=1).fill = _FILL_TOTAL

            _ajuster_largeurs(ws)
            wb.save(chemin_sortie)
            return True
        except Exception as exc:
            logger.error("ExcelBenevoles.generer: %s", exc)
            return False



# ── Style commun Excel Phase 9 ────────────────────────────────────────────────

STYLE_EN_TETE_P9 = PatternFill("solid", fgColor="F0F0F0")
STYLE_TOTAL_P9 = PatternFill("solid", fgColor="E0E0E0")
FONT_GRAS_P9 = Font(bold=True)



def _ecrire_entete_colonne_p9(ws, row: int, colonnes: list[str]) -> None:
    """Écrit une ligne d'en-tête Phase 9."""
    for col_idx, nom in enumerate(colonnes, start=1):
        cell = ws.cell(row=row, column=col_idx, value=nom)
        cell.font = FONT_GRAS_P9
        cell.fill = STYLE_EN_TETE_P9
        cell.alignment = _ALIGN_CENTER



def _appliquer_ligne_total_p9(ws, row: int, nb_cols: int) -> None:
    """Applique le style de total Phase 9 sur une ligne complète."""
    for col_idx in range(1, nb_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = FONT_GRAS_P9
        cell.fill = STYLE_TOTAL_P9



class ExcelAdherents:
    """Génère un export Excel des adhérents."""

    def __init__(self, config_asso: "ConfigAsso") -> None:
        self._config = config_asso

    def generer(self, chemin_sortie: str) -> bool:
        """Génère le classeur Excel des adhérents."""
        try:
            from collections import Counter

            from db.models.membres import get_all_membres

            membres = get_all_membres(include_archives=True)

            wb = Workbook()
            ws_membres = wb.active
            ws_membres.title = "Membres"

            headers = ["Nom", "Prénom", "Statut", "Téléphone", "Email", "Date adhésion"]
            _ecrire_entete_colonne_p9(ws_membres, 1, headers)

            for row_idx, membre in enumerate(membres, start=2):
                ws_membres.cell(row=row_idx, column=1, value=membre.get("nom") or "")
                ws_membres.cell(row=row_idx, column=2, value=membre.get("prenom") or "")
                ws_membres.cell(row=row_idx, column=3, value=membre.get("statut") or "")
                ws_membres.cell(row=row_idx, column=4, value=membre.get("telephone") or "")
                ws_membres.cell(row=row_idx, column=5, value=membre.get("email") or "")
                ws_membres.cell(row=row_idx, column=6, value=_formater_date(membre.get("date_adhesion")))

            if ws_membres.max_row >= 1:
                ws_membres.auto_filter.ref = ws_membres.dimensions
            _ajuster_largeurs(ws_membres)

            ws_resume = wb.create_sheet("Résumé")
            _ecrire_entete_colonne_p9(ws_resume, 1, ["Indicateur", "Valeur"])

            statuts = Counter((membre.get("statut") or "Sans statut") for membre in membres)
            nb_total = len(membres)
            nb_actifs = sum(1 for membre in membres if not bool(membre.get("statut_archive")))

            resume_rows = [
                ("Nombre total de membres", nb_total),
                ("Nombre de membres actifs", nb_actifs),
            ]
            for row_idx, (label, value) in enumerate(resume_rows, start=2):
                ws_resume.cell(row=row_idx, column=1, value=label)
                ws_resume.cell(row=row_idx, column=2, value=value)

            row = 5
            _ecrire_entete_colonne_p9(ws_resume, row, ["Statut", "Nombre"])
            row += 1
            for statut, total in sorted(statuts.items(), key=lambda item: item[0].lower()):
                ws_resume.cell(row=row, column=1, value=statut)
                ws_resume.cell(row=row, column=2, value=total)
                row += 1

            _ajuster_largeurs(ws_resume)
            wb.save(chemin_sortie)
            return True
        except Exception as exc:
            logger.error("ExcelAdherents.generer: %s", exc)
            return False


class ExcelTresorerie:
    """Génère un export Excel de la trésorerie."""

    def __init__(self, config_asso: "ConfigAsso", date_debut: str = '', date_fin: str = '') -> None:
        self._config = config_asso
        self._date_debut = date_debut
        self._date_fin = date_fin

    def generer(self, chemin_sortie: str) -> bool:
        """Génère le classeur Excel de trésorerie."""
        try:
            from db.models import tresorerie as treso_model

            # Compatibilité avec plusieurs variantes du modèle trésorerie.
            get_all_operations = getattr(treso_model, 'get_all_operations', None)
            if callable(get_all_operations):
                operations = get_all_operations(self._date_debut, self._date_fin)
            else:
                operations = treso_model.get_operations(
                    date_debut=self._date_debut or None,
                    date_fin=self._date_fin or None,
                )
            comptes = treso_model.get_all_comptes(actif_only=False)

            wb = Workbook()
            ws_ops = wb.active
            ws_ops.title = "Opérations"
            self._construire_operations(ws_ops, operations)

            ws_categories = wb.create_sheet("Par catégorie")
            self._construire_par_categorie(ws_categories, operations)

            ws_comptes = wb.create_sheet("Par compte")
            self._construire_par_compte(ws_comptes, comptes)

            wb.save(chemin_sortie)
            return True
        except Exception as exc:
            logger.error("ExcelTresorerie.generer: %s", exc)
            return False

    def _construire_operations(self, ws, operations: list[dict]) -> None:
        headers = ["Date", "Libellé", "Type", "Catégorie", "Compte", "Montant", "Statut", "Avancé par", "Statut remboursement"]
        _ecrire_entete_colonne_p9(ws, 1, headers)

        total = 0.0
        for row_idx, operation in enumerate(operations, start=2):
            montant_signe = montant_signe_operation(operation)
            total += montant_signe
            avance_par = ""
            membre_nom = operation.get('avance_par_nom') or operation.get('avance_membre_nom') or ''
            membre_prenom = operation.get('avance_par_prenom') or operation.get('avance_membre_prenom') or ''
            if not membre_nom:
                # Fallback : chercher via avance_par_membre_id dans les membres si dispo
                membre_id = operation.get('avance_par_membre_id')
                if membre_id:
                    avance_par = f"Membre #{membre_id}"
            else:
                avance_par = f"{membre_nom} {membre_prenom}".strip()
            ws.cell(row=row_idx, column=1, value=_formater_date(operation.get('date_operation')))
            ws.cell(row=row_idx, column=2, value=operation.get('libelle') or '')
            ws.cell(row=row_idx, column=3, value=operation.get('type_operation') or '')
            ws.cell(row=row_idx, column=4, value=operation.get('categorie_nom') or 'Sans catégorie')
            ws.cell(row=row_idx, column=5, value=operation.get('compte_nom') or '')
            cell_montant = ws.cell(row=row_idx, column=6, value=round(montant_signe, 2))
            cell_montant.number_format = '#,##0.00'
            ws.cell(row=row_idx, column=7, value=operation.get('statut') or '')
            ws.cell(row=row_idx, column=8, value=avance_par)
            ws.cell(row=row_idx, column=9, value=operation.get('remboursement_statut') or '')

        total_row = max(2, ws.max_row + 1)
        ws.cell(row=total_row, column=1, value='Total')
        cell_total = ws.cell(row=total_row, column=6, value=round(total, 2))
        cell_total.number_format = '#,##0.00'
        _appliquer_ligne_total_p9(ws, total_row, len(headers))

        if ws.max_row >= 1:
            ws.auto_filter.ref = f"A1:I{max(1, total_row - 1)}"
        _ajuster_largeurs(ws)

    def _construire_par_categorie(self, ws, operations: list[dict]) -> None:
        from collections import defaultdict

        headers = ["Catégorie", "Recettes", "Dépenses", "Solde"]
        _ecrire_entete_colonne_p9(ws, 1, headers)

        categories: dict[str, dict[str, float]] = defaultdict(lambda: {"recettes": 0.0, "depenses": 0.0})
        for operation in operations:
            categorie = operation.get('categorie_nom') or 'Sans catégorie'
            montant_signe = montant_signe_operation(operation)
            if montant_signe >= 0:
                categories[categorie]['recettes'] += montant_signe
            else:
                categories[categorie]['depenses'] += abs(montant_signe)

        total_recettes = 0.0
        total_depenses = 0.0
        row = 2
        for categorie, montants in sorted(categories.items(), key=lambda item: item[0].lower()):
            recettes = round(montants['recettes'], 2)
            depenses = round(montants['depenses'], 2)
            solde = round(recettes - depenses, 2)
            total_recettes += recettes
            total_depenses += depenses
            ws.cell(row=row, column=1, value=categorie)
            for col_idx, valeur in ((2, recettes), (3, depenses), (4, solde)):
                cell = ws.cell(row=row, column=col_idx, value=valeur)
                cell.number_format = '#,##0.00'
            row += 1

        ws.cell(row=row, column=1, value='Total')
        totaux_colonnes = (
            (2, round(total_recettes, 2)),
            (3, round(total_depenses, 2)),
            (4, round(total_recettes - total_depenses, 2)),
        )
        for col_idx, valeur in totaux_colonnes:
            cell = ws.cell(row=row, column=col_idx, value=valeur)
            cell.number_format = '#,##0.00'
        _appliquer_ligne_total_p9(ws, row, len(headers))
        _ajuster_largeurs(ws)

    def _construire_par_compte(self, ws, comptes: list[dict]) -> None:
        headers = ["Compte", "Type", "Solde"]
        _ecrire_entete_colonne_p9(ws, 1, headers)

        total = 0.0
        for row_idx, compte in enumerate(comptes, start=2):
            solde = round(float(compte.get('solde_actuel') or 0), 2)
            total += solde
            ws.cell(row=row_idx, column=1, value=compte.get('nom') or '')
            ws.cell(row=row_idx, column=2, value=compte.get('type_compte') or '')
            cell_solde = ws.cell(row=row_idx, column=3, value=solde)
            cell_solde.number_format = '#,##0.00'

        total_row = max(2, ws.max_row + 1)
        ws.cell(row=total_row, column=1, value='Total')
        cell_total = ws.cell(row=total_row, column=3, value=round(total, 2))
        cell_total.number_format = '#,##0.00'
        _appliquer_ligne_total_p9(ws, total_row, len(headers))
        _ajuster_largeurs(ws)


class ExcelStock:
    """Génère un export Excel du stock."""

    def __init__(self, config_asso: "ConfigAsso") -> None:
        self._config = config_asso

    def generer(self, chemin_sortie: str) -> bool:
        """Génère le classeur Excel du stock."""
        try:
            from db.models import stock as stock_model

            articles = stock_model.get_all_articles(include_archives=True)
            # Compatibilité avec plusieurs variantes du modèle stock.
            get_mouvements_stock = getattr(stock_model, 'get_mouvements_stock', None)
            if callable(get_mouvements_stock):
                mouvements = get_mouvements_stock()
            else:
                mouvements = stock_model.get_all_mouvements(limit=5000)

            wb = Workbook()
            ws_articles = wb.active
            ws_articles.title = 'Articles'
            self._construire_articles(ws_articles, articles)

            ws_mouvements = wb.create_sheet('Mouvements')
            self._construire_mouvements(ws_mouvements, mouvements)

            wb.save(chemin_sortie)
            return True
        except Exception as exc:
            logger.error('ExcelStock.generer: %s', exc)
            return False

    @staticmethod
    def _statut_stock(article: dict) -> str:
        if bool(article.get('statut_archive')):
            return 'Archivé'
        quantite = float(article.get('quantite') or 0)
        seuil = float(article.get('seuil_alerte') or 0)
        if quantite <= 0:
            return 'Rupture'
        if seuil and quantite <= seuil:
            return 'Alerte'
        return 'OK'

    def _construire_articles(self, ws, articles: list[dict]) -> None:
        headers = ["Désignation", "Catégorie", "Quantité", "Unité", "Seuil", "Statut stock"]
        _ecrire_entete_colonne_p9(ws, 1, headers)

        for row_idx, article in enumerate(articles, start=2):
            ws.cell(row=row_idx, column=1, value=article.get('nom') or '')
            ws.cell(row=row_idx, column=2, value=article.get('categorie_nom') or '')
            ws.cell(row=row_idx, column=3, value=article.get('quantite') or 0)
            ws.cell(row=row_idx, column=4, value=article.get('unite_nom') or '')
            ws.cell(row=row_idx, column=5, value=article.get('seuil_alerte') or 0)
            ws.cell(row=row_idx, column=6, value=self._statut_stock(article))

        if ws.max_row >= 1:
            ws.auto_filter.ref = ws.dimensions
        _ajuster_largeurs(ws)

    def _construire_mouvements(self, ws, mouvements: list[dict]) -> None:
        headers = ["Date", "Article", "Type mouvement", "Quantité", "Fournisseur", "N° facture"]
        _ecrire_entete_colonne_p9(ws, 1, headers)

        for row_idx, mouvement in enumerate(mouvements, start=2):
            ws.cell(row=row_idx, column=1, value=_formater_date(mouvement.get('date')))
            ws.cell(row=row_idx, column=2, value=mouvement.get('article_nom') or mouvement.get('stock_nom') or '')
            ws.cell(row=row_idx, column=3, value=mouvement.get('type') or '')
            ws.cell(row=row_idx, column=4, value=mouvement.get('quantite') or 0)
            ws.cell(row=row_idx, column=5, value=mouvement.get('fournisseur_nom') or '')
            ws.cell(row=row_idx, column=6, value=mouvement.get('numero_facture') or '')

        if ws.max_row >= 1:
            ws.auto_filter.ref = ws.dimensions
        _ajuster_largeurs(ws)
